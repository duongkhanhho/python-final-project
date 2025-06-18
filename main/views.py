from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import get_object_or_404
from django.utils import timezone
from datetime import datetime, date
from decimal import Decimal
from .models import (
    KhuVuc, QuocGia, DiaDiem, CongViec, PhongBan, 
    NhanVien, NguoiPhuThuoc, Attendance, PayrollRecord
)
from .serializers import (
    KhuVucSerializer, QuocGiaSerializer, DiaDiemSerializer,
    CongViecSerializer, PhongBanSerializer, NhanVienSerializer,
    NhanVienDetailSerializer, NguoiPhuThuocSerializer,
    AttendanceSerializer, AttendanceCreateSerializer, PayrollRecordSerializer,
    PayrollCalculationSerializer
)


class KhuVucViewSet(viewsets.ModelViewSet):
    queryset = KhuVuc.objects.all()
    serializer_class = KhuVucSerializer
    permission_classes = [IsAuthenticated]


class QuocGiaViewSet(viewsets.ModelViewSet):
    queryset = QuocGia.objects.all()
    serializer_class = QuocGiaSerializer
    permission_classes = [IsAuthenticated]


class DiaDiemViewSet(viewsets.ModelViewSet):
    queryset = DiaDiem.objects.all()
    serializer_class = DiaDiemSerializer
    permission_classes = [IsAuthenticated]


class CongViecViewSet(viewsets.ModelViewSet):
    queryset = CongViec.objects.all()
    serializer_class = CongViecSerializer
    permission_classes = [IsAuthenticated]


class PhongBanViewSet(viewsets.ModelViewSet):
    queryset = PhongBan.objects.all()
    serializer_class = PhongBanSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'])
    def employees(self, request, pk=None):
        phong_ban = self.get_object()
        nhan_viens = NhanVien.objects.filter(id_phong_ban=phong_ban)
        serializer = NhanVienSerializer(nhan_viens, many=True)
        return Response(serializer.data)


class NhanVienViewSet(viewsets.ModelViewSet):
    queryset = NhanVien.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return NhanVienDetailSerializer
        return NhanVienSerializer

    @action(detail=True, methods=['get'])
    def dependents(self, request, pk=None):
        nhan_vien = self.get_object()
        nguoi_phu_thuoc = NguoiPhuThuoc.objects.filter(id_nhan_vien=nhan_vien)
        serializer = NguoiPhuThuocSerializer(nguoi_phu_thuoc, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def add_dependent(self, request, pk=None):
        nhan_vien = self.get_object()
        serializer = NguoiPhuThuocSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(id_nhan_vien=nhan_vien)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class NguoiPhuThuocViewSet(viewsets.ModelViewSet):
    queryset = NguoiPhuThuoc.objects.all()
    serializer_class = NguoiPhuThuocSerializer
    permission_classes = [IsAuthenticated]


class AttendanceViewSet(viewsets.ModelViewSet):
    queryset = Attendance.objects.all()
    serializer_class = AttendanceSerializer
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return AttendanceCreateSerializer
        return AttendanceSerializer
    
    @action(detail=False, methods=['post'])
    def check_in(self, request):
        """Nhân viên check-in"""
        id_nhan_vien = request.data.get('id_nhan_vien')
        if not id_nhan_vien:
            return Response({'error': 'ID nhân viên là bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Kiểm tra nhân viên có tồn tại không
        try:
            nhan_vien = NhanVien.objects.get(id_nhan_vien=id_nhan_vien)
        except NhanVien.DoesNotExist:
            return Response({'error': 'Nhân viên không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
        
        # Kiểm tra đã check-in hôm nay chưa
        today = date.today()
        attendance, created = Attendance.objects.get_or_create(
            id_nhan_vien=nhan_vien,
            ngay_lam=today,
            defaults={'check_in': timezone.now()}
        )
        
        if not created and attendance.check_in:
            return Response({'error': 'Đã check-in hôm nay rồi'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not attendance.check_in:
            attendance.check_in = timezone.now()
            attendance.save()
        
        serializer = AttendanceSerializer(attendance)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=False, methods=['post'])
    def check_out(self, request):
        """Nhân viên check-out"""
        id_nhan_vien = request.data.get('id_nhan_vien')
        if not id_nhan_vien:
            return Response({'error': 'ID nhân viên là bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Kiểm tra nhân viên có tồn tại không
        try:
            nhan_vien = NhanVien.objects.get(id_nhan_vien=id_nhan_vien)
        except NhanVien.DoesNotExist:
            return Response({'error': 'Nhân viên không tồn tại'}, status=status.HTTP_404_NOT_FOUND)
        
        # Kiểm tra đã check-in hôm nay chưa
        today = date.today()
        try:
            attendance = Attendance.objects.get(id_nhan_vien=nhan_vien, ngay_lam=today)
        except Attendance.DoesNotExist:
            return Response({'error': 'Chưa check-in hôm nay'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not attendance.check_in:
            return Response({'error': 'Chưa check-in hôm nay'}, status=status.HTTP_400_BAD_REQUEST)
        
        if attendance.check_out:
            return Response({'error': 'Đã check-out hôm nay rồi'}, status=status.HTTP_400_BAD_REQUEST)
        
        attendance.check_out = timezone.now()
        attendance.save()  # Sẽ tự động tính gio_lam và ngay_cong
        
        serializer = AttendanceSerializer(attendance)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Xem tổng hợp chấm công theo tháng"""
        id_nhan_vien = request.query_params.get('id_nhan_vien')
        thang = request.query_params.get('thang')  # Format: YYYY-MM
        
        if not thang:
            thang = timezone.now().strftime('%Y-%m')
        
        try:
            year, month = map(int, thang.split('-'))
            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1)
            else:
                end_date = date(year, month + 1, 1)
        except ValueError:
            return Response({'error': 'Định dạng tháng không đúng (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)
        
        queryset = Attendance.objects.filter(ngay_lam__gte=start_date, ngay_lam__lt=end_date)
        
        if id_nhan_vien:
            queryset = queryset.filter(id_nhan_vien_id=id_nhan_vien)
        
        # Tính tổng hợp
        summary_data = {}
        for attendance in queryset:
            nhan_vien_id = attendance.id_nhan_vien.id_nhan_vien
            if nhan_vien_id not in summary_data:
                summary_data[nhan_vien_id] = {
                    'id_nhan_vien': nhan_vien_id,
                    'ho_ten': f"{attendance.id_nhan_vien.ho} {attendance.id_nhan_vien.ten}",
                    'tong_ngay_lam': 0,
                    'tong_gio_lam': 0,
                    'so_ngay_cham_cong': 0
                }
            
            if attendance.ngay_cong:
                summary_data[nhan_vien_id]['tong_ngay_lam'] += float(attendance.ngay_cong)
            if attendance.gio_lam:
                summary_data[nhan_vien_id]['tong_gio_lam'] += float(attendance.gio_lam)
            summary_data[nhan_vien_id]['so_ngay_cham_cong'] += 1
        
        return Response(list(summary_data.values()), status=status.HTTP_200_OK)


class PayrollRecordViewSet(viewsets.ModelViewSet):
    queryset = PayrollRecord.objects.all()
    serializer_class = PayrollRecordSerializer
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['post'])
    def calculate_payroll(self, request):
        """Tính lương theo tháng"""
        serializer = PayrollCalculationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
        thang = serializer.validated_data['thang']
        id_nhan_vien = serializer.validated_data.get('id_nhan_vien')
        
        # Xác định tháng cần tính
        start_date = date(thang.year, thang.month, 1)
        if thang.month == 12:
            end_date = date(thang.year + 1, 1, 1)
        else:
            end_date = date(thang.year, thang.month + 1, 1)
        
        # Lấy danh sách nhân viên cần tính
        if id_nhan_vien:
            nhan_viens = NhanVien.objects.filter(id_nhan_vien=id_nhan_vien)
        else:
            nhan_viens = NhanVien.objects.all()
        
        results = []
        for nhan_vien in nhan_viens:
            # Tính tổng ngày làm trong tháng
            attendances = Attendance.objects.filter(
                id_nhan_vien=nhan_vien,
                ngay_lam__gte=start_date,
                ngay_lam__lt=end_date
            )
            
            tong_ngay_lam = sum(float(att.ngay_cong) for att in attendances if att.ngay_cong)
            
            # Tạo hoặc cập nhật bảng lương
            payroll_record, created = PayrollRecord.objects.get_or_create(
                id_nhan_vien=nhan_vien,
                thang=start_date,
                defaults={'tong_ngay_lam': tong_ngay_lam}
            )
            
            if not created:
                payroll_record.tong_ngay_lam = tong_ngay_lam
                payroll_record.save()
            
            results.append({
                'id_nhan_vien': nhan_vien.id_nhan_vien,
                'ho_ten': f"{nhan_vien.ho} {nhan_vien.ten}",
                'luong_co_ban': float(nhan_vien.luong),
                'tong_ngay_lam': float(tong_ngay_lam),
                'luong_thuc_nhan': float(payroll_record.luong_thuc_nhan),
                'created': created
            })
        
        return Response({
            'message': f'Đã tính lương tháng {thang.strftime("%m/%Y")}',
            'results': results
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def by_department(self, request):
        """Xem lương theo phòng ban"""
        id_phong_ban = request.query_params.get('id_phong_ban')
        thang = request.query_params.get('thang')  # Format: YYYY-MM
        
        if not id_phong_ban:
            return Response({'error': 'ID phòng ban là bắt buộc'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not thang:
            thang = timezone.now().strftime('%Y-%m')
        
        try:
            year, month = map(int, thang.split('-'))
            start_date = date(year, month, 1)
        except ValueError:
            return Response({'error': 'Định dạng tháng không đúng (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)
        
        queryset = PayrollRecord.objects.filter(
            id_nhan_vien__id_phong_ban_id=id_phong_ban,
            thang=start_date
        )
        
        serializer = PayrollRecordSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Xuất báo cáo bảng lương ra file Excel"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        # Lấy tháng từ query params, mặc định là tháng hiện tại
        thang = request.query_params.get('thang')
        if not thang:
            thang = timezone.now().strftime('%Y-%m')
        
        try:
            year, month = map(int, thang.split('-'))
            start_date = date(year, month, 1)
        except ValueError:
            return Response({'error': 'Định dạng tháng không đúng (YYYY-MM)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Lấy dữ liệu bảng lương
        payrolls = PayrollRecord.objects.filter(thang=start_date)
        
        # Tạo workbook và worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = f"Bang luong thang {month}-{year}"
        
        # Định dạng header
        headers = ['STT', 'ID NV', 'Họ và tên', 'Phòng ban', 'Số ngày làm', 'Lương thực nhận']
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Ghi header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Ghi dữ liệu
        for idx, payroll in enumerate(payrolls, 1):
            nv = payroll.id_nhan_vien
            row = [
                idx,
                nv.id_nhan_vien,
                f"{nv.ho} {nv.ten}",
                nv.id_phong_ban.ten_phong_ban if nv.id_phong_ban else '',
                float(payroll.tong_ngay_lam),
                float(payroll.luong_thuc_nhan)
            ]
            
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=idx+1, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
                
                # Format số tiền
                if col == 6:  # Cột lương thực nhận
                    cell.number_format = '#,##0'
        
        # Điều chỉnh độ rộng cột
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        
        # Tạo response với file Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=bang_luong_{month}_{year}.xlsx'
        
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'])
    def send_emails(self, request):
        """Gửi email bảng lương cho nhân viên"""
        from django.core.mail import EmailMessage
        from django.template.loader import render_to_string
        
        # Lấy tháng từ request data
        thang = request.data.get('thang')
        if not thang:
            thang = timezone.now().strftime('%Y-%m')
            
        try:
            year, month = map(int, thang.split('-'))
            start_date = date(year, month, 1)
        except ValueError:
            return Response({'error': 'Định dạng tháng không đúng (YYYY-MM)'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Lấy dữ liệu bảng lương
        payrolls = PayrollRecord.objects.filter(thang=start_date)
        
        # Lưu nội dung email để trả về
        email_contents = []
        
        for payroll in payrolls:
            nv = payroll.id_nhan_vien
            
            # Chuẩn bị context cho template
            context = {
                'ho_ten': f"{nv.ho} {nv.ten}",
                'thang': f"{month:02d}/{year}",
                'tong_ngay_lam': float(payroll.tong_ngay_lam),
                'luong_thuc_nhan': "{:,.0f}".format(float(payroll.luong_thuc_nhan))
            }
            
            # Render template HTML
            html_content = render_to_string('email/payroll.html', context)
            
            # Tạo email
            subject = f'Bảng lương tháng {month:02d}-{year}'
            email = EmailMessage(
                subject=subject,
                body=html_content,
                from_email='hr@company.com',
                to=['hokhanhduong9204@gmail.com'],  # Demo: gửi tất cả về 1 email
            )
            email.content_subtype = 'html'
            
            # Gửi email
            email.send()
            
            # Lưu nội dung để trả về
            email_contents.append({
                'to': 'hokhanhduong9204@gmail.com',
                'subject': subject,
                'employee': context['ho_ten'],
                'content': html_content
            })
            
        return Response({
            'message': f'Đã gửi {len(email_contents)} email bảng lương tháng {month:02d}/{year}',
            'emails': email_contents
        }, status=status.HTTP_200_OK)
