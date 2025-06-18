from django.core.management.base import BaseCommand
from django.utils import timezone
from main.models import Attendance
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from django.utils.timezone import make_aware, localtime
import traceback

class Command(BaseCommand):
    help = 'Xuất dữ liệu chấm công ra file Excel'

    def add_arguments(self, parser):
        parser.add_argument('--from_date', type=str, help='Ngày bắt đầu (dd/mm/yyyy)')
        parser.add_argument('--to_date', type=str, help='Ngày kết thúc (dd/mm/yyyy)')
        parser.add_argument('--output', type=str, help='Đường dẫn file Excel output')

    def handle(self, *args, **options):
        try:
            # Parse ngày tháng
            if options['from_date']:
                from_date = datetime.strptime(options['from_date'], '%d/%m/%Y').date()
            else:
                # Mặc định là ngày hiện tại
                from_date = timezone.now().date()

            if options['to_date']:
                to_date = datetime.strptime(options['to_date'], '%d/%m/%Y').date()
            else:
                # Mặc định là ngày hiện tại
                to_date = timezone.now().date()

            # Lấy dữ liệu từ database
            attendances = Attendance.objects.filter(
                ngay_lam__range=(from_date, to_date)
            ).select_related('id_nhan_vien').order_by('ngay_lam', 'id_nhan_vien__ho', 'id_nhan_vien__ten')

            self.stdout.write(f"Tìm thấy {attendances.count()} bản ghi chấm công")

            # Debug thông tin bản ghi đầu tiên
            if attendances:
                first = attendances[0]
                self.stdout.write(f"Thông tin bản ghi đầu tiên:")
                self.stdout.write(f"- Nhân viên: {first.id_nhan_vien.ho} {first.id_nhan_vien.ten}")
                self.stdout.write(f"- Ngày làm: {first.ngay_lam}")
                self.stdout.write(f"- Check-in: {first.check_in}")
                self.stdout.write(f"- Check-out: {first.check_out}")
                self.stdout.write(f"- Giờ làm: {first.gio_lam}")
                self.stdout.write(f"- Ngày công: {first.ngay_cong}")

            # Tạo workbook và worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Bảng chấm công"

            # Định dạng header
            headers = ['Họ tên nhân viên', 'Ngày chấm công', 'Giờ check-in', 'Giờ check-out', 'Số giờ làm việc', 'Số công']
            header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            # Ghi dữ liệu
            for row, attendance in enumerate(attendances, 2):
                # Họ tên nhân viên
                ws.cell(row=row, column=1).value = f"{attendance.id_nhan_vien.ho} {attendance.id_nhan_vien.ten}"
                
                # Ngày chấm công
                ws.cell(row=row, column=2).value = attendance.ngay_lam.strftime('%d/%m/%Y')
                
                # Giờ check-in
                if attendance.check_in:
                    # Chuyển naive datetime sang aware datetime
                    aware_checkin = make_aware(attendance.check_in)
                    local_checkin = localtime(aware_checkin)
                    ws.cell(row=row, column=3).value = local_checkin.strftime('%H:%M')
                
                # Giờ check-out
                if attendance.check_out:
                    # Chuyển naive datetime sang aware datetime
                    aware_checkout = make_aware(attendance.check_out)
                    local_checkout = localtime(aware_checkout)
                    ws.cell(row=row, column=4).value = local_checkout.strftime('%H:%M')
                
                # Số giờ làm việc
                ws.cell(row=row, column=5).value = attendance.gio_lam
                
                # Số công
                ws.cell(row=row, column=6).value = attendance.ngay_cong

            # Căn chỉnh cột
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            # Lưu file
            output_file = options['output'] or 'cham_cong.xlsx'
            wb.save(output_file)
            self.stdout.write(self.style.SUCCESS(f'Đã xuất dữ liệu ra file {output_file}'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'Lỗi: {str(e)}'))
            self.stdout.write(self.style.ERROR(traceback.format_exc())) 